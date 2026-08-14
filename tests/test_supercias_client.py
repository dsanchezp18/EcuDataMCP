import io
import ssl

import httpx
import openpyxl
import pytest

from helpers import supercias_client
from helpers.cache import TtlCache

_HEADER = (
    "No. FILA", "EXPEDIENTE", "RUC", "NOMBRE", "SITUACIÓN LEGAL",
    "FECHA_CONSTITUCION", "TIPO", "PAÍS", "REGIÓN", "PROVINCIA", "CANTÓN",
    "CIUDAD", "CALLE", "NÚMERO", "INTERSECCIÓN", "BARRIO", "TELÉFONO",
    "REPRESENTANTE", "CARGO", "CAPITAL SUSCRITO", "CIIU NIVEL 1",
    "CIIU NIVEL 6", "ÚLTIMO BALANCE", "PRESENTÓ BALANCE INICIAL",
    "FECHA PRESENTACIÓN BALANCE INICIAL",
)

_ROW_1 = (
    1, "1", "1790013731001", "ACEITES TROPICALES SOCIEDAD ANONIMA ATSA",
    "ACTIVA", "20/07/1951", "ANÓNIMA", "ECUADOR", "SIERRA", "PICHINCHA",
    "QUITO", "QUITO", "VIA QUININDE KM 37", "SN", "SN", "", "022762426",
    "ACOSTA LLERENA JUAN CARLOS", "GERENTE GENERAL", "48.200,00", "A",
    "A0126.01", "2025", "NO APLICA", "NO APLICA",
)
_ROW_2 = (
    2, "2", "1790004724001", "ACERIA DEL ECUADOR CA ADELCA.", "ACTIVA",
    "17/12/1963", "ANÓNIMA", "ECUADOR", "SIERRA", "GUAYAS", "GUAYAQUIL",
    "GUAYAQUIL", "PANAMERICANA NORTE", "S/N", "S/N", "", "023801321",
    "DIRECACERO DIRECCION DE EMPRESAS DEL ACERO S.A.", "PRESIDENTE EJECUTIVO",
    "125.500.000,00", "C", "C2410.25", "2025", "NO APLICA", "NO APLICA",
)


def _build_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("SUPERINTENDENCIA DE COMPAÑÍAS, VALORES Y SEGUROS",))
    ws.append(("DIRECTORIO DE COMPAÑÍAS",))
    ws.append(("No. DE FILAS: 2",))
    ws.append(("FECHA DE ACTUALIZACION: 13/08/2026 00:53:11",))
    ws.append(_HEADER)
    ws.append(_ROW_1)
    ws.append(_ROW_2)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_xlsx_with_duplicate_ruc() -> bytes:
    dup_row = (
        3, "3", "1790013731001", "ACEITES TROPICALES (EXPEDIENTE DUPLICADO)",
        "ACTIVA", "20/07/1951", "ANÓNIMA", "ECUADOR", "SIERRA", "PICHINCHA",
        "QUITO", "QUITO", "VIA QUININDE KM 37", "SN", "SN", "", "022762426",
        "OTRO REPRESENTANTE", "GERENTE GENERAL", "48.200,00", "A",
        "A0126.01", "2025", "NO APLICA", "NO APLICA",
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("SUPERINTENDENCIA DE COMPAÑÍAS, VALORES Y SEGUROS",))
    ws.append(("DIRECTORIO DE COMPAÑÍAS",))
    ws.append(("No. DE FILAS: 3",))
    ws.append(("FECHA DE ACTUALIZACION: 13/08/2026 00:53:11",))
    ws.append(_HEADER)
    ws.append(_ROW_1)
    ws.append(_ROW_2)
    ws.append(dup_row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture(autouse=True)
def _reset_cache():
    supercias_client._companias_cache = TtlCache(ttl_seconds=60)
    supercias_client._ruc_index_state = None
    yield


def test_col_index():
    assert supercias_client._col_index("A1") == 0
    assert supercias_client._col_index("Z12") == 25
    assert supercias_client._col_index("AA1") == 26


def test_normalize_header():
    assert supercias_client._normalize_header("SITUACIÓN LEGAL") == "situacion_legal"
    assert supercias_client._normalize_header("No. FILA") == "no_fila"
    assert supercias_client._normalize_header("CIIU NIVEL 1") == "ciiu_nivel_1"


def test_parse_xlsx_skips_title_rows_and_detects_header():
    fields, rows = supercias_client._parse_xlsx(_build_xlsx())

    assert fields[2] == "ruc"
    assert fields[3] == "nombre"
    assert len(rows) == 2
    assert rows[0][fields.index("ruc")] == "1790013731001"
    assert rows[0][fields.index("nombre")] == "ACEITES TROPICALES SOCIEDAD ANONIMA ATSA"
    assert rows[1][fields.index("provincia")] == "GUAYAS"


def test_parse_xlsx_raises_without_header():
    only_titles = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(25):
        ws.append(("no header here",))
    wb.save(only_titles)

    with pytest.raises(ValueError, match="encabezado"):
        supercias_client._parse_xlsx(only_titles.getvalue())


async def test_search_companias_by_name_and_ruc(httpx_mock):
    httpx_mock.add_response(url=supercias_client._EXCEL_URL, content=_build_xlsx())

    by_name = await supercias_client.search_companias(query="aceria")
    assert by_name["total"] == 1
    assert by_name["companias"][0]["ruc"] == "1790004724001"

    by_ruc = await supercias_client.search_companias(query="1790013731001")
    assert by_ruc["total"] == 1
    assert by_ruc["companias"][0]["nombre"].startswith("ACEITES TROPICALES")


async def test_search_companias_filters_by_provincia_and_situacion(httpx_mock):
    httpx_mock.add_response(url=supercias_client._EXCEL_URL, content=_build_xlsx())

    result = await supercias_client.search_companias(
        provincia="guayas", situacion_legal="activa"
    )
    assert result["total"] == 1
    assert result["companias"][0]["provincia"] == "GUAYAS"


async def test_search_companias_uses_cache_across_calls(httpx_mock):
    httpx_mock.add_response(url=supercias_client._EXCEL_URL, content=_build_xlsx())

    await supercias_client.search_companias(query="aceria")
    # A second call must not trigger another network request (httpx_mock
    # would fail the test on an unexpected/unmatched extra request).
    await supercias_client.search_companias(query="aceites")


async def test_get_compania_by_ruc_found_and_not_found(httpx_mock):
    httpx_mock.add_response(url=supercias_client._EXCEL_URL, content=_build_xlsx())

    found = await supercias_client.get_compania_by_ruc("1790013731001")
    assert found is not None
    assert found["representante"] == "ACOSTA LLERENA JUAN CARLOS"

    missing = await supercias_client.get_compania_by_ruc("0000000000000")
    assert missing is None


def test_parse_xlsx_raises_when_data_row_exceeds_header_width():
    # Simulate a header row whose last cell is blank (dropped from the row's
    # XML entirely) by writing the real header minus its last column, while
    # a data row still has a value in that trailing column.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(("SUPERINTENDENCIA DE COMPAÑÍAS, VALORES Y SEGUROS",))
    ws.append(("DIRECTORIO DE COMPAÑÍAS",))
    ws.append(("No. DE FILAS: 1",))
    ws.append(("FECHA DE ACTUALIZACION: 13/08/2026 00:53:11",))
    ws.append(_HEADER[:-1])
    ws.append(_ROW_1)
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(ValueError, match="incompleto"):
        supercias_client._parse_xlsx(buf.getvalue())


async def test_duplicate_ruc_keeps_last_row_and_logs_warning(httpx_mock, caplog):
    httpx_mock.add_response(
        url=supercias_client._EXCEL_URL, content=_build_xlsx_with_duplicate_ruc()
    )

    with caplog.at_level("WARNING"):
        found = await supercias_client.get_compania_by_ruc("1790013731001")

    assert found is not None
    assert found["nombre"] == "ACEITES TROPICALES (EXPEDIENTE DUPLICADO)"
    assert any("duplicado" in rec.message for rec in caplog.records)


async def test_ruc_index_is_cached_across_lookups(httpx_mock):
    httpx_mock.add_response(url=supercias_client._EXCEL_URL, content=_build_xlsx())

    await supercias_client.get_compania_by_ruc("1790013731001")
    state_after_first = supercias_client._ruc_index_state
    assert state_after_first is not None

    # A second lookup (and an unrelated search) must reuse the same index
    # object rather than rebuilding it, since the underlying rows haven't
    # changed.
    await supercias_client.get_compania_by_ruc("1790004724001")
    await supercias_client.search_companias(query="aceria")
    assert supercias_client._ruc_index_state is state_after_first


async def test_download_full_retries_insecurely_on_cert_failure(monkeypatch):
    # CKAN_INSECURE_TLS defaults to off (post-cert-renewal); opt in explicitly
    # to exercise the retry path itself.
    monkeypatch.setenv("CKAN_INSECURE_TLS", "1")
    calls: list[bool] = []

    async def fake_download_once(url: str, verify: bool = True) -> bytes:
        calls.append(verify)
        if verify:
            exc = httpx.ConnectError("cert failure")
            exc.__context__ = ssl.SSLCertVerificationError("bad cert")
            raise exc
        return b"ok"

    monkeypatch.setattr(supercias_client, "_download_once", fake_download_once)

    result = await supercias_client._download_full(supercias_client._EXCEL_URL)

    assert result == b"ok"
    assert calls == [True, False]


async def test_download_full_does_not_retry_non_cert_connect_errors(monkeypatch):
    async def fake_download_once(url: str, verify: bool = True) -> bytes:
        raise httpx.ConnectError("connection refused")

    monkeypatch.setattr(supercias_client, "_download_once", fake_download_once)

    with pytest.raises(httpx.ConnectError):
        await supercias_client._download_full(supercias_client._EXCEL_URL)
